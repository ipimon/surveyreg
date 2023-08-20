import streamlit as st
import pandas as pd
import re
import os
import datetime
import time
import pyautogui
from openpyxl import load_workbook
from streamlit_modal import Modal


modal = Modal(key="pop", title='<h2 style="color: orange;">Please Wait...😃</h2>')

# Load Excel data
excel_file = 'allbd.xlsx'
excel_data = pd.ExcelFile(excel_file)
div_data = excel_data.parse('div')
dist_data = excel_data.parse('dist')
upz_data = excel_data.parse('upz')

EXCEL_FILE = "registration_data.xlsx"

def is_valid_phn(contact_number):
    return re.match(r'^[0-9]{11}$', contact_number)

def is_valid_email(email):
    pattern = r"^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9.-]+[.]+[a-zA-Z0-9.-]+$"
    return re.match(pattern, email)

def save_to_excel(data):
    df = pd.DataFrame(data)
    if not os.path.exists(EXCEL_FILE):
        df.to_excel(EXCEL_FILE, index=False)
    else:
        book = load_workbook(EXCEL_FILE)
        writer = pd.ExcelWriter(EXCEL_FILE, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        writer.close()

def save_uploaded_file(uploaded_file):
    current_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    file_extension = uploaded_file.name.split(".")[-1]
    unique_filename = f"{uploaded_file.name.split('.')[0]}_{current_time}.{file_extension}"
    
    with open(os.path.join("pdf_folder", unique_filename), "wb") as f:
        f.write(uploaded_file.getbuffer())
        
        
        
def registration_form():
    st.markdown('<h1 style="text-align: center, color: white;">Surveyor Registration Form</h1>',unsafe_allow_html=True)
    st.markdown("***")
 
  
                          
    # Input fields
    full_name = st.text_input("Full Name")
    contact_number = st.text_input("Contact Number")
 # Validating phone Format
    is_phn_valid = is_valid_phn(contact_number)
    if contact_number and not is_phn_valid:
        st.warning("Please enter a valid phone number")

    

    email = st.text_input("Email")
    # Validating Email Format
    is_email_valid = is_valid_email(email)
    if email and not is_email_valid:
        st.warning("Please enter a valid email address.")
 
    
    division_options = [''] + div_data['Division'].tolist()
    selected_division = st.selectbox('Select Division:', division_options)

    if selected_division != 'Select':
        district_options = [''] + dist_data[dist_data['Division'] == selected_division]['District'].tolist()
        selected_district = st.selectbox('Select District:', district_options)

        if selected_district != 'Select':
            upazila_options = [''] + upz_data[upz_data['District'] == selected_district]['Upazila'].tolist()
            selected_upazila = st.selectbox('Select Upazila:', upazila_options)
        else:
            selected_upazila = 'Select'
    else:
        selected_district = 'Select'
        selected_upazila = 'Select'

    education_qualification = st.selectbox("Education Qualification", ["","SSC", "HSC","Degree(Pass)", "BA Hon's","Honours", "BSS","BSC Hon's", "B.Com Hon's","BBA", "Masters","MSS", "M.Com","MBA", "M.Sc","LLM","Diploma in Engineering"])
    age = st.number_input("Age", min_value=15, max_value=80)
    gender = st.selectbox("Gender", ["", "Male", "Female", "Other"])
    marital_status = st.selectbox("Marital Status", ["", "Single", "Married", "Divorced", "Widowed"])
    experience = st.number_input("Experience (Year)",min_value=0, max_value=50)
    expertise = st.selectbox("Expertise", ["", "Social Survey", "Agricultural Survey", "Educational Survey", "Market Survey","Environmental Survey","Technological Survey", "Telephone Survey", "Online Survey", "Health Survey","Event Management Survey", "FGD","KII"])
    smartphone = st.selectbox("Smartphone Availability", ["", "Yes", "No"])
    cv_upload = st.file_uploader("Upload CV (PDF file)", type=["pdf"])
    
 
    
    # Submit button
    if st.button("Register"):
        if (full_name and is_valid_phn(contact_number) and is_valid_email(email) and
            selected_division and selected_district and selected_upazila and education_qualification and
            age and gender and marital_status and experience and expertise and smartphone and cv_upload):
            # Prepare data
            data = {
                "Full Name": [full_name],
                "Contact Number": [contact_number],
                "Email": [email],
                "Division": [selected_division],
                "District": [selected_district],
                "Upazila": [selected_upazila],
                "Education Qualification": [education_qualification],
                "Age": [age],
                "Gender": [gender],
                "Marital Status": [marital_status],
                "Experience (Year)": [experience],
                "Expertise": [expertise],
                "Smartphone Availability":[smartphone],
                "CV Uploaded": [bool(cv_upload)]
            }
            # Save data to Excel
            save_to_excel(data)
            st.session_state = False
            # Save uploaded PDF file to folder
            if cv_upload:
                save_uploaded_file(cv_upload)
            # Show success message
            if st.success:
                with modal.container():
                        st.markdown('<h3 style="color: green;">Registration successful!😃😁😃</h3>',unsafe_allow_html=True)
                        time.sleep(4)
                        pyautogui.hotkey("ctrl", "F5")
                        
            
               
            # Clear fields
          
        else:
            st.warning("Please fill out all fields correctly.😥😫😥")
            

    st.markdown(
        """
        <style>


            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            


           

        h1 {
            font-size: 45px;
        }
        h2 {
            font-size: 35px;
        }
        h3 {
            font-size: 20px;
        }




        </style>
        """,
        unsafe_allow_html=True,
    )

# Run the registration form
if __name__ == "__main__":
    if not os.path.exists("pdf_folder"):
        os.makedirs("pdf_folder")
    registration_form()

st.markdown("")
st.markdown("")
st.markdown("")
st.markdown("")
st.markdown("")
st.markdown("")
st.markdown("***")
"🟠🟢🟡 iP Skills"
    
 
    
    